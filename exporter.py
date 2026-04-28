import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Style constants
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NEW_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
COMPLETED_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

COLUMNS = ['No', 'IDWORKITEM', 'HEADLINE', 'Status', 'Comments', 'Module', 'Owner', 'Days since Opened', 'Tag']


def _write_header(ws, columns=COLUMNS):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def export_vendor_template(system_issues, output_path):
    """Generate standard vendor template from system export data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Issue List'
    _write_header(ws)

    for idx, issue in enumerate(system_issues, 1):
        row = idx + 1
        values = [
            idx,
            issue.get('ID', ''),
            issue.get('Headline', ''),
            'New',
            '',  # Comments - empty for vendor to fill
            '',  # Module
            '',  # Owner
            issue.get('Days since Opened', ''),
            issue.get('Tag', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    _auto_column_width(ws)
    wb.save(output_path)


def export_issue_list(issues, output_path):
    """Export issue list from displayed table data (includes comments)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Issue List'

    headers = ['No', 'ID', 'Headline', 'Status', 'Comments', 'Module', 'Owner', 'Days since Opened', 'Tag']
    _write_header(ws, headers)

    for idx, issue in enumerate(issues, 1):
        row = idx + 1
        comments = issue.get('comments', '') or ''
        if comments == '-':
            comments = ''
        values = [
            idx,
            issue.get('id', ''),
            issue.get('headline', ''),
            issue.get('current_status', ''),
            comments,
            issue.get('module', ''),
            issue.get('owner', ''),
            issue.get('days', ''),
            issue.get('tag', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    _auto_column_width(ws)
    wb.save(output_path)


HIGHLIGHT_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')


def export_postmortem(issues, all_statuses, output_path):
    """Export postmortem report with per-status duration for each issue."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Postmortem'

    # Dynamic headers: fixed columns + one column per status
    fixed_headers = ['No', 'ID', 'Headline', 'Module', 'Owner', 'Tag', 'Current Status', 'First Seen']
    status_headers = list(all_statuses)
    tail_headers = ['Total Days', 'Resolved', 'Reopened']
    all_headers = fixed_headers + status_headers + tail_headers

    _write_header(ws, all_headers)

    # Status columns get a different header color
    status_start_col = len(fixed_headers) + 1
    STATUS_HEADER_FILL = PatternFill(start_color='2D5F8A', end_color='2D5F8A', fill_type='solid')
    for i, s in enumerate(status_headers):
        cell = ws.cell(row=1, column=status_start_col + i)
        cell.fill = STATUS_HEADER_FILL

    for idx, issue in enumerate(issues, 1):
        row = idx + 1
        values = [
            idx,
            issue['id'],
            issue['headline'],
            issue['module'],
            issue['owner'],
            issue['tag'],
            issue['current_status'],
            issue['first_seen'],
        ]
        # Status duration columns
        for s in status_headers:
            days = issue['status_days'].get(s, '')
            values.append(f'{days}d' if days else '')
        # Tail columns
        values.append(f'{issue["total_days"]}d')
        values.append(issue['resolve_count'] or '')
        values.append(issue['reopen_count'] or '')

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Highlight reopened issues
        if issue['reopen_count'] > 0:
            for col_idx in range(1, len(all_headers) + 1):
                ws.cell(row=row, column=col_idx).fill = HIGHLIGHT_FILL

    _auto_column_width(ws)
    wb.save(output_path)


def _auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
