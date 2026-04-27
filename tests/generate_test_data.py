"""Generate test Excel files for Issue Manager testing.

Creates:
  - test_vendor.xlsx : Vendor document (업체 제공 문제 List)
  - test_system.xlsx : System export (시스템에서 받은 문제 List)

Scenario design:
  - 10 issues in vendor, 10 issues in system
  - 6 common (ongoing - in both files, some with status changes)
  - 4 vendor-only (resolved - vendor has but system doesn't)
  - 4 system-only (new - system has but vendor doesn't)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

today = datetime.now()


def styled_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def styled_row(ws, row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)]
        if not cells:
            continue
        max_len = max((len(str(c.value or '')) for c in cells), default=10)
        ws.column_dimensions[cells[0].column_letter].width = min(max_len + 4, 50)


# ============================================================
# Vendor Document (업체 제공)
# Columns: No, IDWORKITEM, HEADLINE, Status, Comments, Module, Owner, Days since Opened, Tag
# ============================================================
def create_vendor():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Issue List'

    headers = ['No', 'IDWORKITEM', 'HEADLINE', 'Status', 'Comments', 'Module', 'Owner', 'Days since Opened', 'Tag']
    styled_header(ws, headers)

    # 6 common issues (will match system file)
    vendor_data = [
        [1, 'DF260401-00100', 'Audio noise during Dolby Atmos playback on HDMI ARC',           'Problem Analysis', 'Investigating codec path',    'Audio',   'Kim',  15, 'P1'],
        [2, 'DF260401-00101', 'WiFi disconnects intermittently on 5GHz band',                   'Code Review',      'Fix ready for review',        'Network', 'Lee',  22, 'P2'],
        [3, 'DF260401-00102', 'Bluetooth remote pairing fails after firmware update',           'Reproducing',      'Need specific FW version',    'BT',      'Park', 10, 'P1'],
        [4, 'DF260401-00103', 'HDR10+ metadata not passed through to display engine',           'Problem Analysis', 'Checking HAL layer',          'Video',   'Choi',  8, 'P1'],
        [5, 'DF260401-00104', 'CEC power on command ignored from soundbar',                     'New',              '',                            'CEC',     'Jung', 30, 'P3'],
        [6, 'DF260401-00105', 'Subtitle rendering offset in PIP mode',                          'Problem Analysis', 'Offset calculation issue',    'UI',      'Yoon',  5, 'P2'],
        # 4 vendor-only (resolved - not in system anymore)
        [7, 'DF260401-00200', 'Netflix app crash on 4K content launch',                         'Closed',           'Fixed in v2.1.3',            'App',     'Kim',  45, 'P1'],
        [8, 'DF260401-00201', 'HDMI handshake failure with PS5',                                'Closed',           'EDID table updated',          'HDMI',    'Lee',  38, 'P2'],
        [9, 'DF260401-00202', 'Memory leak in media player service',                            'Closed',           'Garbage collection fixed',    'System',  'Park', 60, 'P1'],
        [10,'DF260401-00203', 'IR remote volume key stuck event',                               'Closed',           'Debounce logic added',        'Input',   'Choi', 25, 'P3'],
    ]

    for i, row in enumerate(vendor_data, 2):
        styled_row(ws, i, row)

    auto_width(ws)
    wb.save('test_vendor.xlsx')
    print('Created: test_vendor.xlsx (10 issues: 6 common + 4 resolved)')


# ============================================================
# System Export (시스템에서 받은)
# Row 1: title row (ignored by parser, SYSTEM_HEADER_ROW=1 means header is row 2)
# Row 2: headers
# Columns: IDWORKITEM, HEADLINE, Status, Tag, Opened Time, Seriousness, Frequency, Work Assignment, State Owner, Model Code
# ============================================================
def create_system():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Export'

    # Row 1: title row (parser skips this, reads header from row 2)
    ws.cell(row=1, column=1, value='System Export Report')
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)

    # Row 2: actual headers
    headers = ['IDWORKITEM', 'HEADLINE', 'Status', 'Tag', 'Opened Time', 'Seriousness', 'Frequency', 'Work Assignment', 'State Owner', 'Model Code']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    def days_ago(n):
        return (today - timedelta(days=n)).strftime('%Y-%m-%d %H:%M:%S')

    # 6 common issues (match vendor file - some with status changes!)
    system_data = [
        ['DF260401-00100', 'Audio noise during Dolby Atmos playback on HDMI ARC',           'Code Review',      'P1', days_ago(15), 'Major',    'Always',    'Audio Team',    'Kim',  'QN90D'],
        ['DF260401-00101', 'WiFi disconnects intermittently on 5GHz band',                   'Code Review',      'P2', days_ago(22), 'Major',    'Frequently','Network Team',  'Lee',  'QN85D'],
        ['DF260401-00102', 'Bluetooth remote pairing fails after firmware update',           'Problem Analysis', 'P1', days_ago(10), 'Critical', 'Always',    'BT Team',       'Park', 'QN90D'],
        ['DF260401-00103', 'HDR10+ metadata not passed through to display engine',           'Problem Analysis', 'P1', days_ago(8),  'Major',    'Always',    'Video Team',    'Choi', 'QN95D'],
        ['DF260401-00104', 'CEC power on command ignored from soundbar',                     'Reproducing',      'P3', days_ago(30), 'Minor',    'Sometimes', 'CEC Team',      'Jung', 'QN85D'],
        ['DF260401-00105', 'Subtitle rendering offset in PIP mode',                          'Problem Analysis', 'P2', days_ago(5),  'Minor',    'Always',    'UI Team',       'Yoon', 'QN90D'],
        # 4 system-only (new issues - not in vendor file)
        ['DF260401-00300', 'Screen flicker on Game Mode with VRR enabled',                   'New',              'P1', days_ago(2),  'Critical', 'Sometimes', 'Display Team',  'Han',  'QN95D'],
        ['DF260401-00301', 'Voice assistant activation delay > 3 seconds',                   'Reproducing',      'P2', days_ago(3),  'Major',    'Frequently','AI Team',       'Kang', 'QN90D'],
        ['DF260401-00302', 'USB HDD recording fails on encrypted channels',                  'New',              'P1', days_ago(1),  'Major',    'Always',    'PVR Team',      'Shin', 'QN85D'],
        ['DF260401-00303', 'AirPlay 2 screen mirroring color space mismatch',                'Problem Analysis', 'P2', days_ago(4),  'Minor',    'Always',    'Connectivity',  'Lim',  'QN95D'],
    ]

    for i, row in enumerate(system_data, 3):
        styled_row(ws, i, row)

    auto_width(ws)
    wb.save('test_system.xlsx')
    print('Created: test_system.xlsx (10 issues: 6 common + 4 new)')


if __name__ == '__main__':
    create_vendor()
    create_system()
    print()
    print('Test scenario:')
    print('  Common (ongoing):    6 issues  (DF260401-00100 ~ 00105)')
    print('  Vendor-only (resolved): 4 issues  (DF260401-00200 ~ 00203)')
    print('  System-only (new):   4 issues  (DF260401-00300 ~ 00303)')
    print()
    print('Status changes to detect:')
    print('  00100: Problem Analysis -> Code Review  (vendor says PA, system says CR)')
    print('  00102: Reproducing -> Problem Analysis   (vendor says Repro, system says PA)')
    print('  00104: New -> Reproducing                (vendor says New, system says Repro)')
